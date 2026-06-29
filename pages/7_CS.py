"""
7_CS.py — Customer Success: Activos y Bajas (Monday.com)
Fuente: tablero "Clientes y ex-clientes" (board 6967792411)
  - Grupo "Clienty"  → Activos
  - Resto de grupos  → Bajas
"""
import sys, os, re
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import json
import io
import urllib.request
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Customer Success", page_icon="👥", layout="wide", initial_sidebar_state="expanded")

# ── Configuración Monday ──────────────────────────────────────────

_ID_LTV = "1TGVc9zgYc0siaouIOi8xTOiFopgXuW8AXIB_dqYZ7Ps"

_MONDAY_TOKEN = (
    "eyJhbGciOiJIUzI1NiJ9.eyJ0aWQiOjY2Nzg4MzU5OSwiYWFpIjoxMSwidWlkIjo3N"
    "DQ5MjMwMiwiaWFkIjoiMjAyNi0wNi0wN1QxNzozMTowMi4wMDBaIiwicGVyIjoibWU6"
    "d3JpdGUiLCJhY3RpZCI6MjQxNjExNjcsInJnbiI6InVzZTEifQ.L41MQVmopJ880Q2m"
    "uX6S6erxUv23uOSvppD9fmsoaMQ"
)
_BOARD_ID      = "6967792411"
_GROUP_ACTIVOS = "grupo_nuevo28466"
_COLS          = ["id8__1", "fecha5", "fecha1", "status0", "rubro_mkmttagz", "pain__1"]
_ID_BBDD_CS    = "1pCQtjCZZOrhP21K-EyFECtoNeNNosZfOEgDp9YUZE6M"

_MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


# ── Helpers generales ─────────────────────────────────────────────

def _monday_request(query: str) -> dict:
    payload = json.dumps({"query": query}).encode("utf-8")
    req = urllib.request.Request(
        "https://api.monday.com/v2",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "Authorization": _MONDAY_TOKEN,
            "API-Version":   "2024-01",
        },
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _inferir_fecha_grupo(titulo: str) -> pd.Timestamp:
    t = titulo.lower()
    for mes_str, mes_num in _MESES_ES.items():
        if mes_str in t:
            m = re.search(r"\b(20\d{2})\b", titulo)
            if m:
                return pd.Timestamp(year=int(m.group(1)), month=mes_num, day=1)
    return pd.NaT


def _fmt_fecha(ts) -> str:
    return ts.strftime("%d/%m/%Y") if pd.notna(ts) else "–"


def _norm_id_cs(v) -> str:
    """Normaliza ID CRM: elimina decimales tipo '12345.0'."""
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        s = s[:-2]
    return s if s not in ("nan", "None", "") else ""


def _cat(m):
    """Categoriza meses activos (sin redondear) en intervalos de duración."""
    if pd.isna(m): return None
    m = float(m)
    if m <= 0:  return "0"
    if m <= 3:  return "0-3"
    if m <= 6:  return "+3"
    if m <= 12: return "+6"
    return "+12"


# ── LTV lookup ───────────────────────────────────────────────────

def _leer_sheet_cs(url: str) -> pd.DataFrame:
    import io
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        contenido = resp.read().decode("utf-8")
    df = pd.read_csv(io.StringIO(contenido), dtype=str)
    df.columns = df.columns.str.strip()
    return df


@st.cache_data(ttl=3600)
def _cargar_ltv_lookup_cs() -> dict:
    """Retorna {id_crm: ltv_total}."""
    def _norm(v):
        s = str(v).strip()
        if s in ("", "nan", "None"): return ""
        try: return str(int(float(s)))
        except: return ""

    real_rec = real_impl = prom_rec = pd.Series(dtype=float)

    try:
        df_r = _leer_sheet_cs(f"https://docs.google.com/spreadsheets/d/{_ID_LTV}/export?format=csv&sheet=LTV%20Real")
        _id  = next((c for c in df_r.columns if c.upper() == "ID CRM"), None)
        _usd = next((c for c in df_r.columns if "SECUNDARIA" in c.upper()), None)
        def _parse_usd(s):
            return pd.to_numeric(
                s.astype(str).str.replace(",", ".", regex=False), errors="coerce"
            ).fillna(0)
        df_r["_id"]  = df_r[_id].apply(_norm) if _id else ""
        df_r["_usd"] = _parse_usd(df_r[_usd]) if _usd else 0
        df_r["_impl"] = df_r.get("Producto", pd.Series(dtype=str)).astype(str).str.lower().str.contains("implementa", na=False)
        df_r = df_r[df_r["_id"] != ""]
        real_rec  = df_r[~df_r["_impl"]].groupby("_id")["_usd"].sum()
        real_impl = df_r[df_r["_impl"]].groupby("_id")["_usd"].sum()
    except Exception:
        pass

    try:
        df_p = _leer_sheet_cs(f"https://docs.google.com/spreadsheets/d/{_ID_LTV}/export?format=csv&sheet=LTV%20Prom%20-%202024.08")
        _id_p  = next((c for c in df_p.columns if c.upper() == "ID CRM"), None)
        _usd_p = next((c for c in df_p.columns if "SECUNDARIA" in c.upper()), None)
        df_p["_id"]  = df_p[_id_p].apply(_norm) if _id_p else ""
        df_p["_usd"] = _parse_usd(df_p[_usd_p]) if _usd_p else 0
        prom_rec = df_p[df_p["_id"] != ""].groupby("_id")["_usd"].sum()
    except Exception:
        pass

    # IDs de clientes con venta ANTES de Finnegans (Sep 2024) → prom_rec es legítimo
    _FINNEGANS = pd.Timestamp("2024-09-01")
    _pre_finnegans_ids: set = set()
    try:
        _ID_BBDD = "1pCQtjCZZOrhP21K-EyFECtoNeNNosZfOEgDp9YUZE6M"
        _url_v = f"https://docs.google.com/spreadsheets/d/{_ID_BBDD}/gviz/tq?tqx=out:csv&sheet=BBDD_Ventas"
        df_v = pd.read_csv(_url_v, dtype=str)
        df_v.columns = df_v.columns.str.strip()
        df_v["_fecha"] = pd.to_datetime(df_v.get("Fecha", ""), dayfirst=True, errors="coerce")
        df_v["_id"]    = df_v.get("ID prospecto", pd.Series(dtype=str)).apply(_norm)
        _pre_finnegans_ids = set(
            df_v.loc[df_v["_fecha"] < _FINNEGANS, "_id"].dropna().unique()
        )
    except Exception:
        pass

    lookup = {}
    for idk in set(real_rec.index) | set(real_impl.index) | set(prom_rec.index):
        if not idk: continue
        ltv_real = float(real_rec.get(idk, 0)) + float(real_impl.get(idk, 0))
        # prom_rec solo para clientes con venta pre-Finnegans — evita doble conteo
        # para clientes nuevos que tienen entradas erróneas en LTV Prom
        ltv_prom = float(prom_rec.get(idk, 0)) if idk in _pre_finnegans_ids else 0.0
        lookup[idk] = ltv_real + ltv_prom
    return lookup




# ── Carga Monday ──────────────────────────────────────────────────

@st.cache_data(ttl=86400)
def _descubrir_col_b2b() -> str:
    """Retorna el ID de la columna B2B-B2C en Monday (query liviana)."""
    q = f'{{ boards(ids: [{_BOARD_ID}]) {{ columns {{ id title }} }} }}'
    r = _monday_request(q)
    cols = r["data"]["boards"][0]["columns"]
    for c in cols:
        if "b2b" in c["title"].lower() or "b2c" in c["title"].lower():
            return c["id"]
    return ""


@st.cache_data(ttl=3600)
def cargar_monday_cs() -> pd.DataFrame:
    _col_b2b  = _descubrir_col_b2b()
    _cols_use = list(_COLS) + ([_col_b2b] if _col_b2b else [])
    cols_gql  = ", ".join(f'"{c}"' for c in _cols_use)
    item_fragment = f"""
      id
      name
      group {{ id title }}
      column_values(ids: [{cols_gql}]) {{
        id
        text
      }}
    """
    q = f"""
    {{
      boards(ids: [{_BOARD_ID}]) {{
        items_page(limit: 500) {{
          cursor
          items {{ {item_fragment} }}
        }}
      }}
    }}
    """
    r      = _monday_request(q)
    page   = r["data"]["boards"][0]["items_page"]
    items  = list(page["items"])
    cursor = page.get("cursor")

    while cursor:
        q2 = f"""
        {{
          next_items_page(limit: 500, cursor: "{cursor}") {{
            cursor
            items {{ {item_fragment} }}
          }}
        }}
        """
        r2     = _monday_request(q2)
        page   = r2["data"]["next_items_page"]
        items.extend(page["items"])
        cursor = page.get("cursor")

    rows = []
    for item in items:
        cv = {c["id"]: (c["text"] or "") for c in item["column_values"]}
        rows.append({
            "monday_id":       item["id"],
            "ID CRM":          cv.get("id8__1", ""),
            "Nombre":          item["name"],
            "grupo_id":        item["group"]["id"],
            "grupo_titulo":    item["group"]["title"],
            "_fecha_ingreso":  cv.get("fecha5", ""),
            "_fecha_baja":     cv.get("fecha1", ""),
            "Tipo de cliente": cv.get("status0", ""),
            "Rubro":           cv.get("rubro_mkmttagz", ""),
            "B2B - B2C":       cv.get(_col_b2b, "") if _col_b2b else "",
        })

    df = pd.DataFrame(rows)
    df["_fecha_ingreso"] = pd.to_datetime(df["_fecha_ingreso"], errors="coerce")
    df["_fecha_baja"]    = pd.to_datetime(df["_fecha_baja"],    errors="coerce")
    return df


# ── Carga BBDD_Ventas (para cohortes de venta) ────────────────────

@st.cache_data(ttl=3600)
def cargar_bbdd_ventas_cs() -> pd.DataFrame:
    url = (
        f"https://docs.google.com/spreadsheets/d/{_ID_BBDD_CS}"
        f"/gviz/tq?tqx=out:csv&sheet=BBDD_Ventas"
    )
    df = pd.read_csv(url, dtype=str)
    df.columns = df.columns.str.strip()
    df["_fecha"] = pd.to_datetime(df.get("Fecha", ""), dayfirst=True, errors="coerce")
    df["_id"]    = df.get("ID prospecto", pd.Series(dtype=str)).apply(_norm_id_cs)
    df = df.dropna(subset=["_fecha"])
    df = df[df["_id"] != ""]
    # Un cliente = una cohorte: primera venta
    df = df.sort_values("_fecha").drop_duplicates("_id", keep="first")
    return df.reset_index(drop=True)


# ── CSS compartido para tablas de cohorte ────────────────────────

_COHORTE_CSS = """
* { box-sizing: border-box; margin: 0; }
body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #fff;
    padding: 4px 2px;
}
.wrap {
    overflow: auto;
    max-height: 520px;
    border: 1px solid #c8ccd0;
    border-radius: 4px;
}
table {
    border-collapse: collapse;
    min-width: 860px;
    width: 100%;
    font-size: 0.78rem;
    table-layout: fixed;
}
thead th {
    position: sticky;
    z-index: 2;
    border: 1px dotted #9ab;
}
thead tr:first-child  th { top: 0px; }
thead tr:nth-child(2) th { top: 34px; }

.h-n { background:#1a3a5c; color:#fff; font-weight:bold; padding:6px 4px; text-align:center; }
.h-p { background:#7b341e; color:#fff; font-weight:bold; padding:6px 4px; text-align:center; }
.h-a { background:#145a32; color:#fff; font-weight:bold; padding:6px 4px; text-align:center; }

.sh   { padding:4px 4px; font-weight:bold; text-align:center; border:2px solid #8aabb8; }
.sh-n { background:#a9cce3; }
.sh-p { background:#f5cba7; }
.sh-a { background:#a9dfbf; }

td { border:1px dotted #bbb; padding:4px 4px; text-align:center; }
.c-l { text-align:center; white-space:nowrap; }
.c-n { background:#d6eaf8; }
.c-p { background:#fdebd0; color:#333; }
.c-a { background:#d5f5e3; color:#333; }

.yr-row td   { font-weight:bold; }
.yr-row .c-l { background:#dde1e6; }

.p-row .c-l  { padding-left:16px !important; }

.tbtn {
    background:none; border:none; cursor:pointer;
    font-size:0.65rem; padding:0 3px 0 0;
    color:#444; vertical-align:middle;
}
.tbtn:hover { color:#0055cc; }

td[onclick] { cursor: pointer; }
td[onclick]:hover { filter: brightness(0.82); outline: 1px solid rgba(0,0,0,0.25); }
"""

# JS tabla 1 (bajas por fecha de baja)  — filterClick se embebe en cada HTML
_COHORTE_JS = """
function tog(yr) {
    var rows = document.querySelectorAll('.yr-' + yr);
    var btn  = document.getElementById('btn-' + yr);
    var open = btn.dataset.open === '1';
    rows.forEach(function(r) { r.style.display = open ? 'none' : ''; });
    btn.dataset.open = open ? '0' : '1';
    btn.innerHTML    = open ? '&#9654;' : '&#9660;';
}
"""

# JS tabla 2 (bajas por cohorte de venta)  — filterClick se embebe en cada HTML
_COHORTE_JS_V = """
function togv(yr) {
    var rows = document.querySelectorAll('.yrv-' + yr);
    var btn  = document.getElementById('btnv-' + yr);
    var open = btn.dataset.open === '1';
    rows.forEach(function(r) { r.style.display = open ? 'none' : ''; });
    btn.dataset.open = open ? '0' : '1';
    btn.innerHTML    = open ? '&#9654;' : '&#9660;';
}
"""


# ── Detalle in-iframe ─────────────────────────────────────────────

_DET_DIV = (
    "<div id='det' onclick='closeDet()'"
    " style='display:none;position:fixed;inset:0;background:rgba(0,0,0,0.45);z-index:100;padding:18px 10px;overflow-y:auto'>"
    "<div onclick='event.stopPropagation()'"
    " style='background:#fff;border-radius:6px;max-width:860px;margin:0 auto;"
    "box-shadow:0 4px 24px rgba(0,0,0,0.25)'>"
    "<div style='display:flex;justify-content:space-between;align-items:center;"
    "padding:9px 14px;background:#1a3a5c;border-radius:6px 6px 0 0'>"
    "<span style='color:#fff;font-size:0.82rem;font-weight:bold'>Detalle: <span id='det-title'></span></span>"
    "<button onclick='closeDet()' style='background:#c0392b;color:#fff;border:none;"
    "border-radius:4px;padding:3px 10px;cursor:pointer;font-size:0.72rem'>&#10005; Cerrar</button>"
    "</div>"
    "<div id='det-body' style='max-height:72vh;overflow-y:auto'></div>"
    "</div>"
    "</div>"
)

_IFRAME_JS_TMPL = """
var _SKEYS=['fis','fbs','mr'];
var _lastRows=[], _sortIdx=-1, _sortDir=1;
function _ivl(m){
  if(m==null)return null;
  if(m<=0)return'0';if(m<=3)return'0-3';if(m<=6)return'+3';if(m<=12)return'+6';
  return'+12';
}
function _arr(i){
  return _sortIdx===i?(_sortDir>0?' [^]':' [v]'):' [-]';
}
function filterClick(p){
  var rows=BDATA.filter(function(r){
    if(r.__YR__!==p.yr)return false;
    if(p.months&&p.months.indexOf(r.__MF__)<0)return false;
    if(p.tipo==='ventas')return true;
    if(r.eb===false)return false;
    if(p.iv&&_ivl(r.mr)!==p.iv)return false;
    return true;
  });
  _sortIdx=-1;_sortDir=1;
  renderDet(rows);
}
function closeDet(){document.getElementById('det').style.display='none';}
document.addEventListener('keydown',function(e){if(e.key==='Escape')closeDet();});
function sortDet(i){
  if(_sortIdx===i){_sortDir*=-1;}else{_sortIdx=i;_sortDir=1;}
  var key=_SKEYS[i];
  var s=_lastRows.slice().sort(function(a,b){
    var av=a[key],bv=b[key];
    var an=(av==null||av===''),bn=(bv==null||bv==='');
    if(an&&bn)return 0;if(an)return 1;if(bn)return -1;
    return(av<bv?-1:av>bv?1:0)*_sortDir;
  });
  renderDet(s);
}
function renderDet(rows){
  _lastRows=rows;
  document.getElementById('det-title').textContent=rows.length+' registros';
  var bdy=document.getElementById('det-body');
  if(!rows.length){
    bdy.innerHTML='<p style="padding:12px;color:#666;font-size:0.78rem">Sin registros.</p>';
  } else {
    var S='padding:5px 8px;font-weight:600;background:#1a3a5c;white-space:nowrap;position:sticky;top:0;';
    var h='<table style="width:100%;border-collapse:collapse;font-size:0.78rem">'
      +'<thead><tr style="color:#fff">'
      +'<th style="'+S+'text-align:left">ID CRM</th>'
      +'<th style="'+S+'text-align:left">Nombre</th>'
      +'<th style="'+S+'text-align:center;cursor:pointer" onclick="sortDet(0)">F. ingreso'+_arr(0)+'</th>'
      +'<th style="'+S+'text-align:center;cursor:pointer" onclick="sortDet(1)">F. baja'+_arr(1)+'</th>'
      +'<th style="'+S+'text-align:center;cursor:pointer" onclick="sortDet(2)">Meses'+_arr(2)+'</th>'
      +'<th style="'+S+'text-align:right;cursor:pointer" onclick="sortDet(3)">LTV T'+_arr(3)+'</th>'
      +'</tr></thead><tbody>';
    for(var i=0;i<rows.length;i++){
      var r=rows[i],bg=i%2?'#fff':'#f8f9fa';
      var ltvStr=r.ltv!=null?'$'+r.ltv.toLocaleString('es-AR'):'-';
      h+='<tr style="background:'+bg+'">'
        +'<td style="padding:4px 8px;border-bottom:1px solid #eee">'+(r.id||'-')+'</td>'
        +'<td style="padding:4px 8px;border-bottom:1px solid #eee">'+r.nom+'</td>'
        +'<td style="padding:4px 8px;border-bottom:1px solid #eee;text-align:center">'+r.fi+'</td>'
        +'<td style="padding:4px 8px;border-bottom:1px solid #eee;text-align:center">'+r.fb+'</td>'
        +'<td style="padding:4px 8px;border-bottom:1px solid #eee;text-align:center">'+(r.mr!=null?Math.round(r.mr*10)/10+'m':'-')+'</td>'
        +'<td style="padding:4px 8px;border-bottom:1px solid #eee;text-align:right">'+ltvStr+'</td>'
        +'</tr>';
    }
    bdy.innerHTML=h+'</tbody></table>';
  }
  document.getElementById('det').style.display='block';
}
"""


def _make_iframe_js(bdata_json: str, yr_field: str, m_field: str) -> str:
    """JS con BDATA + filterClick/renderDet embebido en el iframe.
    Escapa </ para que ningún valor rompa el tag <script>."""
    safe_json = bdata_json.replace("</", "<\\/")
    return (
        "var BDATA=" + safe_json + ";\n"
        + _IFRAME_JS_TMPL
        .replace("__YR__", yr_field)
        .replace("__MF__", m_field)
    )


# ── Exportador Excel de cohortes ──────────────────────────────────

def _export_cohorte_excel(df_b: pd.DataFrame, df_v: pd.DataFrame,
                           df_act: pd.DataFrame, gran1: str, gran2: str) -> bytes:
    IVLS  = ["0", "0-3", "+3", "+6", "+12"]
    HDR1  = (26, 84, 92)   # azul oscuro
    HDR2  = (30, 92, 58)   # verde oscuro
    HDR3  = (44, 62, 80)   # azul % acum
    BG0   = (214, 234, 248)
    BG1   = (238, 246, 252)
    BGS   = (169, 204, 227)
    BG0V  = (215, 240, 227)
    BG1V  = (238, 249, 243)
    BGSV  = (163, 217, 189)

    def _hex(rgb): return "%02X%02X%02X" % rgb
    def _fill(rgb): return PatternFill("solid", fgColor=_hex(rgb))
    def _font(bold=False, color="000000", sz=9):
        return Font(name="Arial", bold=bold, color=color, size=sz)
    def _border():
        s = Side(style="thin", color="D0D7DE")
        return Border(left=s, right=s, top=s, bottom=s)
    def _align(h="center"): return Alignment(horizontal=h, vertical="center", wrap_text=False)

    wb = openpyxl.Workbook()

    # ── Hoja 1: Cohorte de Baja ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "Cohorte Baja"

    df = df_b.dropna(subset=["_fecha_baja", "_fecha_ingreso"]).copy()
    df["_y"] = df["_fecha_baja"].dt.year
    df["_m"] = df["_fecha_baja"].dt.month
    df["_meses_r"] = (df["_fecha_baja"] - df["_fecha_ingreso"]).dt.days / 30
    df["_iv"] = df["_meses_r"].apply(_cat)
    df = df.dropna(subset=["_iv"])
    df = df[df["_y"] >= 2024]

    is_trim1 = gran1 == "Trimestre"
    if is_trim1:
        df["_p"] = ((df["_m"] - 1) // 3 + 1)
        def fmt_p1(p): return f"Q{int(p)}"
    else:
        df["_p"] = df["_m"]
        def fmt_p1(p): return str(int(p))

    piv    = df.groupby(["_y", "_p", "_iv"]).size().unstack(fill_value=0).reindex(columns=IVLS, fill_value=0)
    yr_piv = df.groupby(["_y",       "_iv"]).size().unstack(fill_value=0).reindex(columns=IVLS, fill_value=0)
    years  = sorted(df["_y"].unique())

    # Encabezados — fila única para evitar conflictos con celdas fusionadas
    hdrs1 = ["Cohorte", "Total"] + [f"n {iv}" for iv in IVLS] + [f"% {iv}" for iv in IVLS] + [f"acum {iv}" for iv in IVLS]
    fills1 = [HDR1, HDR1] + [HDR1]*5 + [HDR2]*5 + [HDR3]*5
    for i, (lbl, bg) in enumerate(zip(hdrs1, fills1), 1):
        c = ws1.cell(1, i, lbl)
        c.font = _font(bold=True, color="FFFFFF", sz=9)
        c.fill = _fill(bg)
        c.alignment = _align("left" if i == 1 else "center")
        c.border = _border()

    row = 2
    for yr in years:
        yr_cnt = {iv: int(yr_piv.loc[yr, iv]) if yr in yr_piv.index else 0 for iv in IVLS}
        tot    = sum(yr_cnt.values())
        run    = 0.0
        vals   = [str(yr), tot if tot else ""]
        for iv in IVLS: vals.append(yr_cnt.get(iv,0) or "")
        for iv in IVLS:
            n = yr_cnt.get(iv,0); frc = n/tot if tot else 0
            vals.append(f"{round(frc*100)}%" if n else "")
        for iv in IVLS:
            n = yr_cnt.get(iv,0); run += n/tot if tot else 0
            vals.append(f"{round(run*100)}%" if run else "")
        for ci, v in enumerate(vals, 1):
            c = ws1.cell(row, ci, v)
            c.font = _font(bold=True, sz=9)
            c.fill = _fill(BGS)
            c.alignment = _align("left" if ci == 1 else "center")
            c.border = _border()
        row += 1

        if yr in piv.index.get_level_values(0):
            for p in sorted(piv.loc[yr].index):
                p_cnt = {iv: int(piv.loc[(yr,p), iv]) for iv in IVLS}
                tot_p = sum(p_cnt.values()); run_p = 0.0
                is_even = (row % 2 == 0)
                bg = BG0 if is_even else BG1
                vals_p = [f"{yr} {fmt_p1(p)}", tot_p if tot_p else ""]
                for iv in IVLS: vals_p.append(p_cnt.get(iv,0) or "")
                for iv in IVLS:
                    n = p_cnt.get(iv,0); frc = n/tot_p if tot_p else 0
                    vals_p.append(f"{round(frc*100)}%" if n else "")
                for iv in IVLS:
                    n = p_cnt.get(iv,0); run_p += n/tot_p if tot_p else 0
                    vals_p.append(f"{round(run_p*100)}%" if run_p else "")
                for ci, v in enumerate(vals_p, 1):
                    c = ws1.cell(row, ci, v)
                    c.font = _font(sz=9)
                    c.fill = _fill(bg)
                    c.alignment = _align("left" if ci == 1 else "center")
                    c.border = _border()
                row += 1

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 8
    for i in range(3, 18): ws1.column_dimensions[get_column_letter(i)].width = 9
    ws1.row_dimensions[1].height = 18; ws1.row_dimensions[2].height = 16

    # ── Hoja 2: Cohorte de Venta ──────────────────────────────────
    ws2 = wb.create_sheet("Cohorte Venta")

    df_vtas = df_v[df_v["_fecha"].dt.year >= 2024].copy()
    df_vtas["_y"] = df_vtas["_fecha"].dt.year
    df_vtas["_m"] = df_vtas["_fecha"].dt.month
    is_trim2 = gran2 == "Trimestre"
    if is_trim2:
        df_vtas["_p"] = ((df_vtas["_m"] - 1) // 3 + 1)
        def fmt_p2(p): return f"Q{int(p)}"
    else:
        df_vtas["_p"] = df_vtas["_m"]
        def fmt_p2(p): return str(int(p))

    _tmp = (df_b[df_b["_fecha_baja"].notna() & df_b["_fecha_ingreso"].notna() & (df_b["ID CRM"] != "")]
            .drop_duplicates("ID CRM").copy())
    _tmp["_meses_raw"] = (_tmp["_fecha_baja"] - _tmp["_fecha_ingreso"]).dt.days / 30
    baja_map2 = _tmp.set_index("ID CRM")["_meses_raw"].to_dict()
    df_vtas["_meses_activos"] = df_vtas["_id"].map(baja_map2)
    df_vtas["_iv"] = df_vtas["_meses_activos"].apply(_cat)

    ventas_piv2 = df_vtas.groupby(["_y","_p"]).size()
    ventas_yr2  = df_vtas.groupby("_y").size()
    df_baj2     = df_vtas.dropna(subset=["_iv"]).copy()
    bajas_piv2  = df_baj2.groupby(["_y","_p"]).size()
    bajas_yr2   = df_baj2.groupby("_y").size()
    bajas_iv_piv2 = (df_baj2.groupby(["_y","_p","_iv"]).size()
                     .unstack(fill_value=0).reindex(columns=IVLS, fill_value=0))
    bajas_iv_yr2  = (df_baj2.groupby(["_y","_iv"]).size()
                     .unstack(fill_value=0).reindex(columns=IVLS, fill_value=0))
    years2 = sorted(df_vtas["_y"].unique())

    # Encabezados hoja 2 — fila única
    hdrs2 = ["Cohorte","Ventas","Bajas"] + [f"n {iv}" for iv in IVLS] + [f"% {iv}" for iv in IVLS] + [f"acum {iv}" for iv in IVLS]
    fills2 = [HDR1, HDR1, HDR1] + [HDR2]*5 + [HDR1]*5 + [HDR3]*5
    for i, (lbl, bg) in enumerate(zip(hdrs2, fills2), 1):
        c = ws2.cell(1, i, lbl)
        c.font = _font(bold=True, color="FFFFFF", sz=9)
        c.fill = _fill(bg)
        c.alignment = _align("left" if i == 1 else "center")
        c.border = _border()

    row2 = 2
    for yr in years2:
        vtas_yr = int(ventas_yr2.get(yr, 0))
        baj_yr  = int(bajas_yr2.get(yr, 0))
        yr_iv   = {iv: int(bajas_iv_yr2.loc[yr, iv]) if yr in bajas_iv_yr2.index else 0 for iv in IVLS}
        run = 0.0
        vals = [str(yr), vtas_yr or "", baj_yr or ""]
        for iv in IVLS: vals.append(yr_iv.get(iv,0) or "")
        for iv in IVLS:
            n = yr_iv.get(iv,0); frc = n/vtas_yr if vtas_yr else 0
            vals.append(f"{round(frc*100)}%" if n else "")
        for iv in IVLS:
            n = yr_iv.get(iv,0); run += n/vtas_yr if vtas_yr else 0
            vals.append(f"{round(run*100)}%" if run else "")
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row2, ci, v)
            c.font = _font(bold=True, sz=9); c.fill = _fill(BGSV)
            c.alignment = _align("left" if ci == 1 else "center"); c.border = _border()
        row2 += 1

        periods_yr = sorted(set(df_vtas[df_vtas["_y"]==yr]["_p"].unique()))
        for p in periods_yr:
            vtas_p = int(ventas_piv2.get((yr,p), 0))
            baj_p  = int(bajas_piv2.get((yr,p), 0))
            p_iv   = ({iv: int(bajas_iv_piv2.loc[(yr,p), iv]) for iv in IVLS}
                      if (yr,p) in bajas_iv_piv2.index else {iv: 0 for iv in IVLS})
            run_p = 0.0; is_even = (row2 % 2 == 0); bg = BG0V if is_even else BG1V
            vals_p = [f"{yr} {fmt_p2(p)}", vtas_p or "", baj_p or ""]
            for iv in IVLS: vals_p.append(p_iv.get(iv,0) or "")
            for iv in IVLS:
                n = p_iv.get(iv,0); frc = n/vtas_p if vtas_p else 0
                vals_p.append(f"{round(frc*100)}%" if n else "")
            for iv in IVLS:
                n = p_iv.get(iv,0); run_p += n/vtas_p if vtas_p else 0
                vals_p.append(f"{round(run_p*100)}%" if run_p else "")
            for ci, v in enumerate(vals_p, 1):
                c = ws2.cell(row2, ci, v)
                c.font = _font(sz=9); c.fill = _fill(bg)
                c.alignment = _align("left" if ci == 1 else "center"); c.border = _border()
            row2 += 1

    ws2.column_dimensions["A"].width = 14
    for i in range(2, 19): ws2.column_dimensions[get_column_letter(i)].width = 9
    ws2.row_dimensions[1].height = 18; ws2.row_dimensions[2].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Tabla 1: cohorte de baja × intervalo de duración ─────────────

def _cohorte_html(df_b: pd.DataFrame, granularity: str) -> tuple[str, int]:
    IVLS = ["0", "0-3", "+3", "+6", "+12"]

    df = df_b.dropna(subset=["_fecha_baja", "_fecha_ingreso"]).copy()
    df["_y"]       = df["_fecha_baja"].dt.year
    df["_m"]       = df["_fecha_baja"].dt.month
    df["_meses_r"] = (df["_fecha_baja"] - df["_fecha_ingreso"]).dt.days / 30
    df["_iv"]      = df["_meses_r"].apply(_cat)
    df = df.dropna(subset=["_iv"])
    df = df[df["_y"] >= 2022]

    if df.empty:
        return "<p style='padding:12px;color:#666'>Sin datos desde 2022.</p>", 60

    if granularity == "trimestre":
        df["_p"] = ((df["_m"] - 1) // 3 + 1)
        def fmt_p(p): return f"Q{int(p)}"
    else:
        df["_p"] = df["_m"]
        def fmt_p(p): return str(int(p))

    piv = (
        df.groupby(["_y", "_p", "_iv"])
        .size().unstack(fill_value=0)
        .reindex(columns=IVLS, fill_value=0)
    )
    yr_piv = (
        df.groupby(["_y", "_iv"])
        .size().unstack(fill_value=0)
        .reindex(columns=IVLS, fill_value=0)
    )
    years = sorted(df["_y"].unique())

    # ── Datos para detalle in-iframe ──────────────────────────
    _bdata_rows: list[dict] = []
    for _, _r in df.iterrows():
        _fi_ok  = pd.notna(_r["_fecha_ingreso"])
        _fb_ok  = pd.notna(_r["_fecha_baja"])
        _id_str = _norm_id_cs(str(_r.get("ID CRM") or ""))
        _ltv_v = _ltv_lookup_cs.get(_id_str)
        _bdata_rows.append({
            "id":  _id_str,
            "nom": str(_r.get("Nombre") or ""),
            "fi":  _fmt_fecha(_r["_fecha_ingreso"]) if _fi_ok else "–",
            "fb":  _fmt_fecha(_r["_fecha_baja"])    if _fb_ok else "–",
            "fis": _r["_fecha_ingreso"].strftime("%Y-%m-%d") if _fi_ok else "",
            "fbs": _r["_fecha_baja"].strftime("%Y-%m-%d")    if _fb_ok else "",
            "mr":  round(float(_r["_meses_r"]), 1)  if pd.notna(_r["_meses_r"]) else None,
            "yr":  int(_r["_y"]),
            "mb":  int(_r["_m"]),
            "ltv": round(_ltv_v) if _ltv_v is not None else None,
        })
    _bdata_json = json.dumps(_bdata_rows, ensure_ascii=False)
    _filter_js  = _make_iframe_js(_bdata_json, "yr", "mb")
    # ─────────────────────────────────────────────────────────

    def _cells(cnt, fb):
        total = sum(cnt.get(iv, 0) for iv in IVLS)
        n_html = p_html = a_html = ""
        running = 0.0
        for iv in IVLS:
            n   = cnt.get(iv, 0)
            frc = n / total if total > 0 else 0.0
            running += frc
            if n > 0:
                _fj = json.dumps({**fb, "iv": iv})
                oc  = f" onclick='filterClick({_fj})'"
            else:
                oc = ""
            n_html += f'<td class="c-n"{oc}>{n if n > 0 else ""}</td>'
            p_html += f'<td class="c-p">{str(round(frc * 100)) + "%" if n > 0 else ""}</td>'
            a_html += f'<td class="c-a">{str(round(running * 100)) + "%" if running > 0 else ""}</td>'
        return total, n_html, p_html, a_html

    tbody    = ""
    n_p_rows = 0

    for yr in years:
        fb_yr  = {"src": "baja", "yr": int(yr), "months": None, "iv": None}
        yr_cnt = {iv: int(yr_piv.loc[yr, iv]) if yr in yr_piv.index else 0 for iv in IVLS}
        tot, nh, ph, ah = _cells(yr_cnt, fb_yr)
        if tot > 0:
            _fj_yr = json.dumps(fb_yr)
            oc_yr  = f" onclick='filterClick({_fj_yr})'"
        else:
            oc_yr = ""
        tbody += (
            f'<tr class="yr-row">'
            f'<td class="c-l"><button class="tbtn" id="btn-{yr}" data-open="0"'
            f' onclick="tog(\'{yr}\')">&#9654;</button>{yr}</td>'
            f'<td class="c-n"{oc_yr}>{tot if tot > 0 else ""}</td>'
            f'{nh}{ph}{ah}</tr>'
        )
        if yr in piv.index.get_level_values(0):
            for p in sorted(piv.loc[yr].index):
                n_p_rows += 1
                if granularity == "trimestre":
                    q        = int(p)
                    months_p = list(range((q - 1) * 3 + 1, q * 3 + 1))
                else:
                    months_p = [int(p)]
                fb_p  = {"src": "baja", "yr": int(yr), "months": months_p, "iv": None}
                p_cnt = {iv: int(piv.loc[(yr, p), iv]) for iv in IVLS}
                tot_p, nh_p, ph_p, ah_p = _cells(p_cnt, fb_p)
                if tot_p > 0:
                    _fj_p = json.dumps(fb_p)
                    oc_p  = f" onclick='filterClick({_fj_p})'"
                else:
                    oc_p = ""
                tbody += (
                    f'<tr class="p-row yr-{yr}" style="display:none">'
                    f'<td class="c-l">{yr}&nbsp;{fmt_p(p)}</td>'
                    f'<td class="c-n"{oc_p}>{tot_p if tot_p > 0 else ""}</td>'
                    f'{nh_p}{ph_p}{ah_p}</tr>'
                )

    sub_n = "".join(f'<th class="sh sh-n">{iv}</th>' for iv in IVLS)
    sub_p = "".join(f'<th class="sh sh-p">{iv}</th>' for iv in IVLS)
    sub_a = "".join(f'<th class="sh sh-a">{iv}</th>' for iv in IVLS)
    height = min(max(15 * 28 + 90, (len(years) + n_p_rows) * 28 + 90), 600)

    html = (
        "<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<style>{_COHORTE_CSS}</style></head><body>"
        "<div class='wrap'><table><thead>"
        "<tr>"
        "<th rowspan='2' class='h-n'>Cohorte</th>"
        "<th rowspan='2' class='h-n'>Total</th>"
        "<th colspan='5' class='h-n'>Bajas por fecha de baja</th>"
        "<th colspan='5' class='h-p'>%</th>"
        "<th colspan='5' class='h-a'>% Acum</th>"
        "</tr>"
        f"<tr>{sub_n}{sub_p}{sub_a}</tr>"
        "</thead>"
        f"<tbody>{tbody}</tbody>"
        "</table></div>"
        + _DET_DIV
        + f"<script>{_COHORTE_JS}</script>"
        + f"<script>{_filter_js}</script>"
        + "</body></html>"
    )
    return html, height


# ── Tabla 2: cohorte de VENTA × bajas (Monday) ───────────────────

def _cohorte_ventas_html(
    df_b: pd.DataFrame,
    df_v: pd.DataFrame,
    df_act: pd.DataFrame,
    granularity: str,
) -> tuple[str, int]:
    """
    Tabla de cohorts por MES DE VENTA (BBDD_Ventas).
    - Ventas: cantidad de ventas de ese cohorte
    - Bajas: cuántas de esas ventas están hoy dadas de baja en Monday
    - Intervalos: desglose de bajas por Meses activos
    - % sobre ventas: cada columna / Ventas del cohorte
    """
    IVLS = ["0", "0-3", "+3", "+6", "+12"]

    # Preparar ventas desde 2024
    df_vtas = df_v.copy()
    df_vtas = df_vtas[df_vtas["_fecha"].dt.year >= 2024].copy()
    if df_vtas.empty:
        return "<p style='padding:12px;color:#666'>Sin datos de ventas desde 2024.</p>", 60

    df_vtas["_y"] = df_vtas["_fecha"].dt.year
    df_vtas["_m"] = df_vtas["_fecha"].dt.month

    if granularity == "trimestre":
        df_vtas["_p"] = ((df_vtas["_m"] - 1) // 3 + 1)
        def fmt_p(p): return f"Q{int(p)}"
    else:
        df_vtas["_p"] = df_vtas["_m"]
        def fmt_p(p): return str(int(p))

    # Mapa CRM ID → meses activos SIN redondear (para categorizar correctamente los límites)
    _tmp = (
        df_b[df_b["_fecha_baja"].notna() & df_b["_fecha_ingreso"].notna() & (df_b["ID CRM"] != "")]
        .drop_duplicates("ID CRM")
        .copy()
    )
    _tmp["_meses_raw"] = (_tmp["_fecha_baja"] - _tmp["_fecha_ingreso"]).dt.days / 30
    _tmp_idx = _tmp.set_index("ID CRM")
    baja_map     = _tmp_idx["_meses_raw"].to_dict()
    nom_map      = _tmp_idx["Nombre"].to_dict()
    fi_map       = {k: _fmt_fecha(v) for k, v in _tmp_idx["_fecha_ingreso"].to_dict().items()}
    fb_map       = {k: _fmt_fecha(v) for k, v in _tmp_idx["_fecha_baja"].to_dict().items()}
    fi_sort_map  = {k: v.strftime("%Y-%m-%d") if pd.notna(v) else ""
                    for k, v in _tmp_idx["_fecha_ingreso"].to_dict().items()}
    fb_sort_map  = {k: v.strftime("%Y-%m-%d") if pd.notna(v) else ""
                    for k, v in _tmp_idx["_fecha_baja"].to_dict().items()}

    # Extender mapas con activos (para clientes que no son bajas)
    _hoy_v = pd.Timestamp(date.today())
    _act_clean = (
        df_act[df_act["_fecha_ingreso"].notna() & (df_act["ID CRM"] != "")]
        .drop_duplicates("ID CRM")
        .set_index("ID CRM")
    )
    act_mr_map: dict = {}
    for _aid, _arow in _act_clean.iterrows():
        if _aid not in nom_map:
            nom_map[_aid] = str(_arow.get("Nombre") or "")
        if _aid not in fi_map:
            fi_map[_aid]      = _fmt_fecha(_arow["_fecha_ingreso"])
            fi_sort_map[_aid] = _arow["_fecha_ingreso"].strftime("%Y-%m-%d")
        # Meses activos calculado a hoy (se actualiza cada vez que se carga)
        act_mr_map[_aid] = round((_hoy_v - _arow["_fecha_ingreso"]).days / 30, 1)

    # Enriquecer ventas con info de baja
    df_vtas["_meses_activos"] = df_vtas["_id"].map(baja_map)
    df_vtas["_iv"]            = df_vtas["_meses_activos"].apply(_cat)

    # Pivots de ventas
    ventas_piv = df_vtas.groupby(["_y", "_p"]).size()
    ventas_yr  = df_vtas.groupby("_y").size()

    # Pivots de bajas (solo filas con iv válido)
    df_baj = df_vtas.dropna(subset=["_iv"]).copy()
    bajas_piv    = df_baj.groupby(["_y", "_p"]).size()
    bajas_yr     = df_baj.groupby("_y").size()
    bajas_iv_piv = (
        df_baj.groupby(["_y", "_p", "_iv"])
        .size().unstack(fill_value=0)
        .reindex(columns=IVLS, fill_value=0)
    )
    bajas_iv_yr = (
        df_baj.groupby(["_y", "_iv"])
        .size().unstack(fill_value=0)
        .reindex(columns=IVLS, fill_value=0)
    )

    # ── Datos para detalle in-iframe (todos los clientes del cohorte) ─
    _baj_ids = set(df_baj["_id"].unique())
    _bdata_rows_v: list[dict] = []
    for _, _r in df_vtas.iterrows():
        _vid    = str(_r.get("_id") or "")
        _is_baj = _vid in _baj_ids
        _mr_ok  = _is_baj and pd.notna(_r.get("_meses_activos"))
        _bdata_rows_v.append({
            "id":  _vid,
            "nom": str(nom_map.get(_vid, _vid) or ""),
            "fi":  fi_map.get(_vid, "–"),
            "fis": fi_sort_map.get(_vid, ""),
            "fb":  fb_map.get(_vid, "–") if _is_baj else "Activo",
            "fbs": fb_sort_map.get(_vid, "") if _is_baj else "",
            "mr":  (round(float(_r["_meses_activos"]), 1) if _mr_ok
                    else act_mr_map.get(_vid)),
            "yv":  int(_r["_y"]),
            "mv":  int(_r["_m"]),
            "eb":  _is_baj,
        })
    _bdata_json_v = json.dumps(_bdata_rows_v, ensure_ascii=False)
    _filter_js_v  = _make_iframe_js(_bdata_json_v, "yv", "mv")
    # ─────────────────────────────────────────────────────────

    years = sorted(df_vtas["_y"].unique())

    def _cells(ventas, bajas, cnt, fb):
        # Conteos de intervalos (con onclick)
        n_html = ""
        for iv in IVLS:
            n = cnt.get(iv, 0)
            if n > 0 and fb is not None:
                _fj = json.dumps({**fb, "iv": iv})
                oc  = f" onclick='filterClick({_fj})'"
            else:
                oc = ""
            n_html += f'<td class="c-n"{oc}>{n if n > 0 else ""}</td>'

        # % sobre ventas
        p_html  = ""
        tot_pct = str(round(bajas / ventas * 100)) + "%" if ventas > 0 and bajas > 0 else ""
        for iv in IVLS:
            n   = cnt.get(iv, 0)
            pct = str(round(n / ventas * 100)) + "%" if ventas > 0 and n > 0 else ""
            p_html += f'<td class="c-p">{pct}</td>'

        # % acumulado
        a_html  = ""
        running = 0
        for iv in IVLS:
            running += cnt.get(iv, 0)
            pct = str(round(running / ventas * 100)) + "%" if ventas > 0 and running > 0 else ""
            a_html += f'<td class="c-a">{pct}</td>'

        return n_html, tot_pct, p_html, a_html

    tbody    = ""
    n_p_rows = 0

    for yr in years:
        fb_yr  = {"src": "venta", "yr": int(yr), "months": None, "iv": None}
        ven_yr = int(ventas_yr.get(yr, 0))
        baj_yr = int(bajas_yr.get(yr, 0))
        cnt_yr = {
            iv: int(bajas_iv_yr.loc[yr, iv])
            if yr in bajas_iv_yr.index else 0
            for iv in IVLS
        }
        nh, tpct, ph, ah = _cells(ven_yr, baj_yr, cnt_yr, fb_yr)
        if ven_yr > 0:
            _fj_ven_yr = json.dumps({**fb_yr, "tipo": "ventas"})
            oc_ven_yr  = f" onclick='filterClick({_fj_ven_yr})'"
        else:
            oc_ven_yr = ""
        if baj_yr > 0:
            _fj_baj = json.dumps(fb_yr)
            oc_baj  = f" onclick='filterClick({_fj_baj})'"
        else:
            oc_baj = ""

        act_yr   = ven_yr - baj_yr
        act_yr_pct = f' <span style="font-size:0.7em;font-weight:normal">({round(act_yr/ven_yr*100)}%)</span>' if ven_yr > 0 else ""
        tbody += (
            f'<tr class="yr-row">'
            f'<td class="c-l"><button class="tbtn" id="btnv-{yr}" data-open="0"'
            f' onclick="togv(\'{yr}\')">&#9654;</button>{yr}</td>'
            f'<td class="c-n"{oc_ven_yr}>{ven_yr if ven_yr > 0 else ""}</td>'
            f'<td class="c-n">{act_yr if act_yr > 0 else ""}{act_yr_pct}</td>'
            f'<td class="c-n"{oc_baj}>{baj_yr if baj_yr > 0 else ""}</td>'
            f'{nh}'
            f'<td class="c-p">{tpct}</td>'
            f'{ph}'
            f'{ah}</tr>'
        )

        if yr in ventas_piv.index.get_level_values(0):
            for p in sorted(ventas_piv.loc[yr].index):
                n_p_rows += 1
                if granularity == "trimestre":
                    q        = int(p)
                    months_p = list(range((q - 1) * 3 + 1, q * 3 + 1))
                else:
                    months_p = [int(p)]
                fb_p  = {"src": "venta", "yr": int(yr), "months": months_p, "iv": None}
                ven_p = int(ventas_piv.get((yr, p), 0))
                baj_p = int(bajas_piv.get((yr, p), 0))
                act_p = ven_p - baj_p
                act_p_pct = f' <span style="font-size:0.7em;font-weight:normal">({round(act_p/ven_p*100)}%)</span>' if ven_p > 0 else ""
                cnt_p = {}
                if yr in bajas_iv_piv.index.get_level_values(0):
                    try:
                        cnt_p = {iv: int(bajas_iv_piv.loc[(yr, p), iv]) for iv in IVLS}
                    except KeyError:
                        cnt_p = {iv: 0 for iv in IVLS}
                nh_p, tpct_p, ph_p, ah_p = _cells(ven_p, baj_p, cnt_p, fb_p)
                if ven_p > 0:
                    _fj_ven_p = json.dumps({**fb_p, "tipo": "ventas"})
                    oc_ven_p  = f" onclick='filterClick({_fj_ven_p})'"
                else:
                    oc_ven_p = ""
                if baj_p > 0:
                    _fj_bp = json.dumps(fb_p)
                    oc_bp  = f" onclick='filterClick({_fj_bp})'"
                else:
                    oc_bp = ""

                tbody += (
                    f'<tr class="p-row yrv-{yr}" style="display:none">'
                    f'<td class="c-l">{yr}&nbsp;{fmt_p(p)}</td>'
                    f'<td class="c-n"{oc_ven_p}>{ven_p if ven_p > 0 else ""}</td>'
                    f'<td class="c-n">{act_p if act_p > 0 else ""}{act_p_pct}</td>'
                    f'<td class="c-n"{oc_bp}>{baj_p if baj_p > 0 else ""}</td>'
                    f'{nh_p}'
                    f'<td class="c-p">{tpct_p}</td>'
                    f'{ph_p}'
                    f'{ah_p}</tr>'
                )

    # Sub-headers
    sub_n = "".join(f'<th class="sh sh-n">{iv}</th>' for iv in IVLS)
    sub_p = "".join(f'<th class="sh sh-p">{iv}</th>' for iv in IVLS)
    sub_a = "".join(f'<th class="sh sh-a">{iv}</th>' for iv in IVLS)
    height = min(max(15 * 28 + 90, (len(years) + n_p_rows) * 28 + 90), 600)

    html = (
        "<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<style>{_COHORTE_CSS}</style></head><body>"
        "<div class='wrap'><table><thead>"
        "<tr>"
        "<th rowspan='2' class='h-n'>Cohorte</th>"
        "<th rowspan='2' class='h-n'>Ventas</th>"
        "<th rowspan='2' class='h-n'>Activos</th>"
        "<th rowspan='2' class='h-n'>Bajas</th>"
        "<th colspan='5' class='h-n'>Bajas por fecha de ingreso</th>"
        "<th colspan='6' class='h-p'>% sobre ventas</th>"
        "<th colspan='5' class='h-a'>% Acum</th>"
        "</tr>"
        "<tr>"
        f"{sub_n}"
        "<th class='sh sh-p'>Total</th>"
        f"{sub_p}"
        f"{sub_a}"
        "</tr>"
        "</thead>"
        f"<tbody>{tbody}</tbody>"
        "</table></div>"
        + _DET_DIV
        + f"<script>{_COHORTE_JS_V}</script>"
        + f"<script>{_filter_js_v}</script>"
        + "</body></html>"
    )
    return html, height


# ── UI principal ──────────────────────────────────────────────────

st.title("👥 Customer Success")

try:
    df_all = cargar_monday_cs()
except Exception as e:
    st.error(f"Error al conectar con Monday.com: {e}")
    st.stop()

try:
    _ltv_lookup_cs = _cargar_ltv_lookup_cs()
except Exception:
    _ltv_lookup_cs = {}


if df_all.empty:
    st.warning("Sin datos en Monday.")
    st.stop()

# 1. Deduplicar por monday_id
df_all = df_all.drop_duplicates(subset=["monday_id"]).copy()

# 2. Separar activos / bajas
df_activos = df_all[df_all["grupo_id"] == _GROUP_ACTIVOS].copy()
df_bajas   = df_all[df_all["grupo_id"] != _GROUP_ACTIVOS].copy()

# 3. Completar fecha_baja desde nombre del grupo cuando está vacía
_mask_sin_baja = df_bajas["_fecha_baja"].isna()
if _mask_sin_baja.any():
    df_bajas.loc[_mask_sin_baja, "_fecha_baja"] = (
        df_bajas.loc[_mask_sin_baja, "grupo_titulo"].apply(_inferir_fecha_grupo)
    )

# 4. Deduplicar bajas por ID CRM
_df_b_con_id = df_bajas[df_bajas["ID CRM"] != ""].copy()
_df_b_sin_id = df_bajas[df_bajas["ID CRM"] == ""].copy()
_df_b_con_id = (
    _df_b_con_id
    .sort_values("_fecha_baja", ascending=False, na_position="last")
    .drop_duplicates(subset=["ID CRM"], keep="first")
)
df_bajas = pd.concat([_df_b_con_id, _df_b_sin_id], ignore_index=True)

# 5. Meses activos
_hoy = pd.Timestamp(date.today())

df_activos["Meses activos"] = (
    (_hoy - df_activos["_fecha_ingreso"]).dt.days / 30
).round(0)

# 5b. LTV Total
def _norm_id_cs_ltv(v):
    s = str(v).strip()
    if s in ("", "nan", "None"): return ""
    try: return str(int(float(s)))
    except: return ""

df_activos["LTV T"] = df_activos["ID CRM"].apply(
    lambda v: _ltv_lookup_cs.get(_norm_id_cs_ltv(v), None)
)

df_bajas["Meses activos"] = (
    (df_bajas["_fecha_baja"] - df_bajas["_fecha_ingreso"]).dt.days / 30
).round(0)

# 6. Fechas formateadas
df_activos["Fecha de ingreso"] = df_activos["_fecha_ingreso"].apply(_fmt_fecha)
df_bajas["Fecha de ingreso"]   = df_bajas["_fecha_ingreso"].apply(_fmt_fecha)
df_bajas["Fecha de baja"]      = df_bajas["_fecha_baja"].apply(_fmt_fecha)


# ── Sidebar ───────────────────────────────────────────────────────

with st.sidebar:
    st.title("👥 Customer Success")
    st.markdown("---")
    if st.button("🔄 Actualizar datos", use_container_width=True,
                 help="Recarga datos de Monday.com y BBDD_Ventas"):
        st.cache_data.clear()
        st.rerun()
    st.caption("Fuente: Monday.com · BBDD_Ventas")


# ── KPI card helper ───────────────────────────────────────────────

def _kpi(lbl, val, sub=""):
    return (
        '<div style="flex:1;min-width:150px;background:white;border:1px solid #e2e8f0;'
        'border-radius:10px;padding:16px 12px;text-align:center;'
        'box-shadow:0 1px 3px rgba(0,0,0,0.06)">'
        f'<div style="font-size:1.45rem;font-weight:700;color:#1e293b">{val}</div>'
        f'<div style="font-size:0.75rem;color:#64748b;margin-top:4px">{lbl}</div>'
        + (f'<div style="font-size:0.68rem;color:#94a3b8;margin-top:2px">{sub}</div>' if sub else "")
        + '</div>'
    )


# ── Tabs ──────────────────────────────────────────────────────────

tab_act, tab_ob, tab_baj = st.tabs(["Resumen", "OB", "📉 Bajas"])


# ════════════════════════════════════════════════════════════════
#  TAB ACTIVOS
# ════════════════════════════════════════════════════════════════

with tab_act:

    _n_act = len(df_activos)

    # ── Distribución por intervalos de meses ─────────────────
    _IVLS_ACT = [
        ("0",   0,  0),
        ("1-3", 1,  3),
        ("+3",  4,  6),
        ("+6",  7, 12),
        ("+12",13, 18),
        ("+18",19, 24),
        ("+24",25, 36),
        ("+36",37, 9999),
    ]
    _ma_vals = df_activos["Meses activos"].dropna()
    _dist_rows = []
    for lbl, lo, hi in _IVLS_ACT:
        n = int(((lo <= _ma_vals) & (_ma_vals <= hi)).sum())
        _dist_rows.append({"Intervalo": lbl, "Clientes": n,
                           "%": f"{round(n / _n_act * 100)}%" if _n_act > 0 else "–"})
    _df_dist = pd.DataFrame(_dist_rows)

    # ── Distribución por rubro (top 7 + Otros) ───────────────
    _rubro_counts = (
        df_activos["Rubro"]
        .fillna("").replace("", "Sin rubro")
        .value_counts()
    )
    _top7  = _rubro_counts.head(7)
    _otros = _rubro_counts.iloc[7:].sum()
    _rubro_rows = [{"Rubro": r, "Clientes": int(n),
                    "%": f"{round(n / _n_act * 100)}%" if _n_act > 0 else "–"}
                   for r, n in _top7.items()]
    if _otros > 0:
        _rubro_rows.append({"Rubro": "Otros", "Clientes": int(_otros),
                             "%": f"{round(_otros / _n_act * 100)}%" if _n_act > 0 else "–"})
    _df_rubro = pd.DataFrame(_rubro_rows)

    def _tabla_html(df, col_izq):
        _TH = ("background:#f1f5f9;font-size:0.75rem;color:#64748b;font-weight:600;"
               "padding:6px 10px;border-bottom:2px solid #e2e8f0;")
        _TD_L = "font-size:0.82rem;padding:5px 10px;border-bottom:1px solid #f1f5f9;text-align:left;"
        _TD_C = "font-size:0.82rem;padding:5px 10px;border-bottom:1px solid #f1f5f9;text-align:center;"
        cols = list(df.columns)
        head = "".join(
            f'<th style="{_TH}text-align:{"left" if c == col_izq else "center"}">{c}</th>'
            for c in cols
        )
        rows = ""
        for _, r in df.iterrows():
            rows += "<tr>" + "".join(
                f'<td style="{_TD_L if c == col_izq else _TD_C}">{r[c]}</td>'
                for c in cols
            ) + "</tr>"
        col_w = f"{100 // len(cols)}%"
        col_tags = "".join(f'<col style="width:{col_w}">' for _ in cols)
        return (
            '<div style="border:1px solid #e2e8f0;border-radius:8px;overflow:hidden">'
            f'<table style="width:100%;border-collapse:collapse;table-layout:fixed">'
            f'<colgroup>{col_tags}</colgroup>'
            f'<thead><tr>{head}</tr></thead>'
            f'<tbody>{rows}</tbody></table></div>'
        )

    _col_kpi, _col_dist, _col_rubro = st.columns([1, 1, 1])
    with _col_kpi:
        st.markdown(
            '<div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:20px">'
            + _kpi("Clientes activos", f"{_n_act:,}")
            + '</div>',
            unsafe_allow_html=True,
        )
    with _col_dist:
        st.markdown(_tabla_html(_df_dist, "Intervalo"), unsafe_allow_html=True)
    with _col_rubro:
        st.markdown(_tabla_html(_df_rubro, "Rubro"), unsafe_allow_html=True)

    # (OB movido a tab_ob)

with tab_ob:
    try:
        import datos_crm as _dcrm
        import json as _json_ob
        _df_ob = _dcrm.cargar_ob_detalle()

        if not _df_ob.empty:
            _ALIAS = {"Melina": "Meli", "julispinelli": "Juli", "Nicolas Guzmán": "Nico", "Nicolas Guzman": "Nico"}
            _df_ob["estratega"] = _df_ob["estratega"].replace(_ALIAS)
            _ETAPA_ALIAS = {
                "OB 3 practica en vivo + Bot": "OB 3 Practica",
                "OB 3 Practica en vivo + Bot": "OB 3 Practica",
            }
            _df_ob["etapa"] = _df_ob["etapa"].str.replace(
                r"OB 3 [Pp]ractica en vivo \+ Bot\)?", "OB3 (Practica)", regex=True
            )
            _df_ob["etapa"] = _df_ob["etapa"].str.replace(
                "Por cerrar imp.", "Por cerrar", regex=False
            )
            _df_ob["etapa"] = _df_ob["etapa"].str.replace(
                r"OB 2 \(secuencias?\)", "OB 2 (sec)", regex=True
            )

            st.markdown('<div style="margin-top:24px"></div>', unsafe_allow_html=True)
            st.markdown("### Onboarding")

            _estrategas = sorted(_df_ob["estratega"].unique())
            _etapas     = sorted(_df_ob["etapa"].unique())
            _ob_json    = _json_ob.dumps(_df_ob.to_dict(orient="records"), ensure_ascii=False)

            # ── helpers de celda ────────────────────────────────
            def _td_num(n, filt, cls_bg, bold=False):
                _b = "font-weight:bold;" if bold else ""
                if n > 0:
                    return (f'<td class="{cls_bg}" data-filter="{filt}"'
                            f' onclick="obModal(this.dataset.filter)"'
                            f' style="border:1px dotted #bbb;padding:4px 6px;text-align:center;'
                            f'cursor:pointer;{_b}">{n}</td>')
                return f'<td class="{cls_bg}" style="border:1px dotted #bbb;padding:4px 6px;text-align:center;{_b}"></td>'

            def _td_lbl(txt, bold=False):
                _b = "font-weight:bold;" if bold else ""
                return (f'<td class="c-n" style="border:1px dotted #bbb;padding:4px 8px;'
                        f'text-align:center;white-space:nowrap;{_b}">{txt}</td>')

            # ── Tabla SLA ────────────────────────────────────────
            _sla_cols = ["≤30d", ">30d", "Sin fecha", "Total"]
            _sla_cls  = {"≤30d": "c-n", ">30d": "c-n", "Sin fecha": "c-n", "Total": "c-n"}

            _thead_sla = (
                '<tr>'
                '<th class="h-n" style="border:1px dotted #9ab;padding:6px 4px;text-align:center"></th>'
                + "".join(
                    f'<th class="h-n" style="border:1px dotted #9ab;padding:6px 4px;text-align:center">{c}</th>'
                    for c in _sla_cols
                )
                + '</tr>'
            )

            def _sla_data_row(est, sub, is_total=False):
                vals = {
                    "≤30d":      int((sub["sla"] == "≤30d").sum()),
                    ">30d":      int((sub["sla"] == ">30d").sum()),
                    "Sin fecha": int((sub["sla"] == "Sin fecha").sum()),
                    "Total":     len(sub),
                }
                row = _td_lbl(est, bold=is_total)
                for col in _sla_cols:
                    filt = f"sla:{col}" if is_total else f"est:{est}|sla:{col}"
                    row += _td_num(vals[col], filt, _sla_cls[col], bold=is_total)
                cls = ' class="yr-row"' if is_total else ""
                return f"<tr{cls}>{row}</tr>"

            _tbody_sla = ""
            for _est in _estrategas:
                _tbody_sla += _sla_data_row(_est, _df_ob[_df_ob["estratega"] == _est])
            _tbody_sla += _sla_data_row("Total", _df_ob, is_total=True)

            # ── Tabla Riesgo ─────────────────────────────────────
            _riesgos = [r for r in ["Alto", "Medio", "Bajo", "—"] if r in _df_ob["riesgo"].values]

            _thead_riesgo = (
                '<tr>'
                '<th class="h-n" style="border:1px dotted #9ab;padding:6px 4px;text-align:center"></th>'
                + "".join(
                    f'<th class="h-n" style="border:1px dotted #9ab;padding:6px 4px;text-align:center">{r}</th>'
                    for r in _riesgos
                )
                + '<th class="h-n" style="border:1px dotted #9ab;padding:6px 4px;text-align:center">Total</th>'
                '</tr>'
            )

            def _riesgo_data_row(est, sub, is_total=False):
                row = _td_lbl(est, bold=is_total)
                for r in _riesgos:
                    n    = int((sub["riesgo"] == r).sum())
                    filt = f"riesgo:{r}" if is_total else f"est:{est}|riesgo:{r}"
                    row += _td_num(n, filt, "c-n", bold=is_total)
                filt_tot = "all" if is_total else f"est:{est}"
                row += _td_num(len(sub), filt_tot, "c-n", bold=is_total)
                cls = ' class="yr-row"' if is_total else ""
                return f"<tr{cls}>{row}</tr>"

            _tbody_riesgo = ""
            for _est in _estrategas:
                _tbody_riesgo += _riesgo_data_row(_est, _df_ob[_df_ob["estratega"] == _est])
            _tbody_riesgo += _riesgo_data_row("Total", _df_ob, is_total=True)

            # ── Tabla Etapa ──────────────────────────────────────
            _thead_etapa = (
                '<tr>'
                '<th class="h-n" style="border:1px dotted #9ab;padding:6px 4px;text-align:center"></th>'
                + "".join(
                    f'<th class="h-n" style="border:1px dotted #9ab;padding:6px 4px;text-align:center">{et}</th>'
                    for et in _etapas
                )
                + '<th class="h-n" style="border:1px dotted #9ab;padding:6px 4px;text-align:center">Total</th>'
                '</tr>'
            )

            def _etapa_data_row(est, sub, is_total=False):
                row = _td_lbl(est, bold=is_total)
                for et in _etapas:
                    n    = int((sub["etapa"] == et).sum())
                    filt = f"etapa:{et}" if is_total else f"est:{est}|etapa:{et}"
                    row += _td_num(n, filt, "c-n", bold=is_total)
                filt_tot = "all" if is_total else f"est:{est}"
                row += _td_num(len(sub), filt_tot, "c-n", bold=is_total)
                cls = ' class="yr-row"' if is_total else ""
                return f"<tr{cls}>{row}</tr>"

            _tbody_etapa = ""
            for _est in _estrategas:
                _tbody_etapa += _etapa_data_row(_est, _df_ob[_df_ob["estratega"] == _est])
            _tbody_etapa += _etapa_data_row("Total", _df_ob, is_total=True)

            _n_ests  = len(_estrategas)
            _ob_height = max((_n_ests + 2) * 34 + 200, 260)

            _n_rows_det = len(_df_ob)
            _ob_height  = ((_n_ests + 2) * 34 + 60) + (_n_rows_det * 30 + 120)

            _html_ob = f"""<!DOCTYPE html><html><head><meta charset='utf-8'>
<style>
*{{box-sizing:border-box;margin:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#fff;padding:4px 0;font-size:0.78rem}}
.layout{{display:flex;gap:20px;align-items:flex-start;margin-bottom:20px}}
.col-sla{{flex:0 0 26%}}
.col-riesgo{{flex:0 0 26%}}
.col-riesgo .tbl-title{{height:3rem}}
.col-etapa{{flex:1;overflow-x:auto}}
.col-sla table, .col-riesgo table, .col-etapa table{{table-layout:fixed;width:100%}}
.tbl-title{{font-size:0.82rem;font-weight:600;color:#1e293b;margin-bottom:5px;height:2.4rem;display:flex;align-items:flex-end}}
.wrap{{border:1px solid #c8ccd0;border-radius:4px;overflow:hidden}}
table{{border-collapse:collapse;width:100%;table-layout:auto}}
.h-n{{background:#1a3a5c;color:#fff;font-weight:bold;padding:6px 4px;text-align:center;border:1px dotted #9ab;font-size:0.72rem}}
.c-n{{background:#d6eaf8}}
.yr-row td{{font-weight:bold}}
td[onclick]:hover{{filter:brightness(0.88);outline:1px solid rgba(0,0,0,0.2);cursor:pointer}}
/* Filtros detalle */
.det-filters{{display:flex;gap:10px;align-items:center;margin-bottom:10px;flex-wrap:wrap}}
.det-filters select{{font-size:0.78rem;padding:5px 10px;border:1px solid #cbd5e1;border-radius:6px;background:#fff;color:#374151;cursor:pointer;min-width:120px}}
.det-count{{font-size:0.78rem;color:#64748b;margin-left:4px}}
/* Tabla detalle */
.det-wrap{{border:1px solid #e2e8f0;border-radius:6px;overflow:hidden}}
.det-tbl{{width:100%;border-collapse:collapse;font-size:0.8rem}}
.det-th{{padding:6px 12px;font-weight:600;background:#f1f5f9;color:#475569;border-bottom:2px solid #e2e8f0;cursor:pointer;white-space:nowrap;text-align:left}}
.det-th:hover{{background:#e2e8f0}}
.det-td{{padding:6px 12px;border-bottom:1px solid #f1f5f9;color:#1e293b}}
.det-td-c{{padding:6px 12px;border-bottom:1px solid #f1f5f9;color:#1e293b;text-align:center}}
tr:hover .det-td, tr:hover .det-td-c{{background:#f8fafc}}
</style>
</head><body>

<div class="layout">
  <div class="col-sla">
    <div class="tbl-title">Por estratega y SLA</div>
    <div class="wrap"><table><thead>{_thead_sla}</thead><tbody>{_tbody_sla}</tbody></table></div>
  </div>
  <div class="col-riesgo">
    <div class="tbl-title">Por estratega y riesgo</div>
    <div class="wrap"><table><thead>{_thead_riesgo}</thead><tbody>{_tbody_riesgo}</tbody></table></div>
  </div>
  <div class="col-etapa">
    <div class="tbl-title">Por estratega y etapa</div>
    <div class="wrap"><table><thead>{_thead_etapa}</thead><tbody>{_tbody_etapa}</tbody></table></div>
  </div>
</div>

<div class="det-filters">
  <select id="f-est"    onchange="applyFilters()"><option value="">Estratega: Todos</option></select>
  <select id="f-riesgo" onchange="applyFilters()"><option value="">Riesgo: Todos</option></select>
  <select id="f-etapa"  onchange="applyFilters()"><option value="">Etapa: Todos</option></select>
  <select id="f-sla"    onchange="applyFilters()"><option value="">SLA: Todos</option></select>
  <span class="det-count" id="ob-count"></span>
</div>
<div class="det-wrap">
  <table class="det-tbl">
    <thead><tr>
      <th class="det-th" onclick="sortOb(0)">Nombre <span id="arr0"></span></th>
      <th class="det-th" onclick="sortOb(1)" style="text-align:center">Estratega <span id="arr1"></span></th>
      <th class="det-th" onclick="sortOb(2)" style="text-align:center">Etapa <span id="arr2"></span></th>
      <th class="det-th" onclick="sortOb(3)" style="text-align:center">Inicio impl. <span id="arr3"></span></th>
      <th class="det-th" onclick="sortOb(4)" style="text-align:center">Días OB <span id="arr4"></span></th>
      <th class="det-th" onclick="sortOb(5)" style="text-align:center">SLA <span id="arr5"></span></th>
      <th class="det-th" onclick="sortOb(6)" style="text-align:center">Riesgo <span id="arr6"></span></th>
    </tr></thead>
    <tbody id="ob-tbl-body"></tbody>
  </table>
</div>

<script>
var obData   = {_ob_json};
var _allRows = obData.slice();
var _curRows = obData.slice();
var _obSort  = -1, _obDir = 1;
var _KEYS    = ['nombre','estratega','etapa','inicio','dias','sla','riesgo'];
var _RC      = {{'Alto':'#c0392b','Medio':'#d68910','Bajo':'#1e8449'}};

function _uniq(key) {{
  var s={{}};_allRows.forEach(function(r){{s[r[key]]=1;}});return Object.keys(s).sort();
}}
function _fill(id, key, label) {{
  var sel=document.getElementById(id);
  _uniq(key).forEach(function(v){{
    var o=document.createElement('option');o.value=v;o.textContent=label+v;sel.appendChild(o);
  }});
}}
_fill('f-est','estratega','Estratega: ');
_fill('f-riesgo','riesgo','Riesgo: ');
_fill('f-etapa','etapa','Etapa: ');
_fill('f-sla','sla','SLA: ');

function applyFilters() {{
  var est=document.getElementById('f-est').value;
  var rie=document.getElementById('f-riesgo').value;
  var eta=document.getElementById('f-etapa').value;
  var sla=document.getElementById('f-sla').value;
  _curRows=_allRows.filter(function(r){{
    return(!est||r.estratega===est)&&(!rie||r.riesgo===rie)
         &&(!eta||r.etapa===eta)&&(!sla||r.sla===sla);
  }});
  _obSort=-1;_obDir=1;renderOb(_curRows);
}}

function filterClick(filter) {{
  [['f-est',''],['f-riesgo',''],['f-etapa',''],['f-sla','']].forEach(function(p){{
    document.getElementById(p[0]).value='';
  }});
  if(filter!=='all') filter.split('|').forEach(function(p){{
    var kv=p.split(':'),key=kv[0],val=kv.slice(1).join(':');
    if(key==='est')   document.getElementById('f-est').value=val;
    if(key==='sla')   document.getElementById('f-sla').value=val;
    if(key==='etapa') document.getElementById('f-etapa').value=val;
    if(key==='riesgo')document.getElementById('f-riesgo').value=val;
  }});
  applyFilters();
}}

function sortOb(i) {{
  if(_obSort===i){{_obDir*=-1;}}else{{_obSort=i;_obDir=1;}}
  var key=_KEYS[i];
  _curRows=_curRows.slice().sort(function(a,b){{
    var av=a[key],bv=b[key];
    var an=(av==null||av==='—'),bn=(bv==null||bv==='—');
    if(an&&bn)return 0;if(an)return 1;if(bn)return-1;
    return(av<bv?-1:av>bv?1:0)*_obDir;
  }});
  renderOb(_curRows);
}}

function renderOb(rows) {{
  document.getElementById('ob-count').textContent=rows.length+' clientes';
  for(var i=0;i<7;i++){{
    document.getElementById('arr'+i).textContent=_obSort===i?(_obDir>0?' ↑':' ↓'):'';
  }}
  var h='';
  rows.forEach(function(r,i){{
    var rc=_RC[r.riesgo]||'#333';
    h+='<tr>'
      +'<td class="det-td">'+r.nombre+'</td>'
      +'<td class="det-td-c">'+r.estratega+'</td>'
      +'<td class="det-td-c">'+r.etapa+'</td>'
      +'<td class="det-td-c">'+r.inicio+'</td>'
      +'<td class="det-td-c">'+r.dias+'</td>'
      +'<td class="det-td-c">'+r.sla+'</td>'
      +'<td class="det-td-c" style="color:'+rc+';font-weight:600">'+r.riesgo+'</td>'
      +'</tr>';
  }});
  if(!rows.length)h='<tr><td colspan="7" style="padding:14px;color:#94a3b8;text-align:center">Sin registros.</td></tr>';
  document.getElementById('ob-tbl-body').innerHTML=h;
}}

// Reemplazar onclick de celdas de resumen por filterClick
document.querySelectorAll('td[data-filter]').forEach(function(td){{
  td.onclick=function(){{filterClick(this.dataset.filter);}};
}});

renderOb(_allRows);
</script>
</body></html>"""

            st.components.v1.html(_html_ob, height=_ob_height, scrolling=True)

    except Exception as _ob_e:
        st.exception(_ob_e)

with tab_act:
    _fc1, _fc2, _fc3 = st.columns([2, 1, 1])
    _buscar_a = _fc1.text_input("🔍 Buscar nombre o ID", "", key="buscar_activos")

    # Filtro Intervalo
    def _ivl_label(m):
        if pd.isna(m): return None
        m = float(m)
        if m <= 0:  return "0"
        if m <= 3:  return "1-3"
        if m <= 6:  return "+3"
        if m <= 12: return "+6"
        if m <= 18: return "+12"
        if m <= 24: return "+18"
        if m <= 36: return "+24"
        return "+36"

    _ivl_opciones = ["0", "1-3", "+3", "+6", "+12", "+18", "+24", "+36"]
    _filtro_ivl   = _fc2.multiselect("Intervalo", _ivl_opciones, key="fil_ivl_act",
                                      placeholder="Todos", label_visibility="visible")

    # Filtro Rubro
    _rubros_opciones = sorted(df_activos["Rubro"].dropna().replace("", pd.NA).dropna().unique().tolist())
    _filtro_rubro    = _fc3.multiselect("Rubro", _rubros_opciones, key="fil_rubro_act",
                                         placeholder="Todos", label_visibility="visible")

    _cols_act = ["ID CRM", "Nombre", "Fecha de ingreso", "Meses activos",
                 "Tipo de cliente", "Rubro", "LTV T"]
    _show_a   = df_activos.sort_values("Meses activos", ascending=False)[_cols_act].rename(columns={
        "ID CRM":          "ID",
        "Fecha de ingreso": "Fecha ing",
        "Meses activos":    "Meses act",
    })

    if _buscar_a:
        _mask = (
            _show_a["Nombre"].str.contains(_buscar_a, case=False, na=False)
            | _show_a["ID"].str.contains(_buscar_a, case=False, na=False)
        )
        _show_a = _show_a[_mask]

    if _filtro_ivl:
        _show_a = _show_a[
            df_activos.loc[_show_a.index, "Meses activos"].apply(_ivl_label).isin(_filtro_ivl)
        ]

    if _filtro_rubro:
        _show_a = _show_a[_show_a["Rubro"].isin(_filtro_rubro)]

    st.markdown(f"**{len(_show_a):,} clientes**")
    st.dataframe(
        _show_a.style.set_properties(
            subset=["Meses act", "LTV T"], **{"text-align": "center"}
        ),
        use_container_width=True, hide_index=True,
        column_config={
            "ID":        st.column_config.TextColumn("ID",        width=None),
            "Nombre":    st.column_config.TextColumn("Nombre",    width="medium"),
            "Fecha ing": st.column_config.TextColumn("Fecha ing", width=None),
            "Meses act": st.column_config.NumberColumn(
                "Meses act", format="%.0f", width="small",
                help="(Hoy − Fecha de ingreso) ÷ 30",
            ),
            "Tipo de cliente": st.column_config.TextColumn("Tipo de cliente", width=None),
            "LTV T": st.column_config.NumberColumn("LTV T", format="$ %.0f", width=None),
        },
    )
    st.markdown(
        '<div style="text-align:right;font-size:0.68rem;color:#94a3b8;margin-top:4px">'
        'Fuente: Monday.com · Grupo "Clienty"</div>',
        unsafe_allow_html=True,
    )


# ════════════════════════════════════════════════════════════════
#  TAB BAJAS
# ════════════════════════════════════════════════════════════════

with tab_baj:

    # ── Cargar ventas ─────────────────────────────────────────
    try:
        df_ventas_cs = cargar_bbdd_ventas_cs()
    except Exception as _e:
        st.error(f"Error cargando BBDD_Ventas: {_e}")
        df_ventas_cs = pd.DataFrame()

    # ── Filtros ───────────────────────────────────────────────
    _fc_r, _fc_y, _fc_b2b = st.columns(3)

    _rubros_baj_opc = ["Todos"] + sorted(
        df_bajas["Rubro"].dropna().replace("", pd.NA).dropna().unique().tolist()
    )
    _filtro_rubro_dist = _fc_r.selectbox("Rubro", _rubros_baj_opc, key="fil_rubro_baj_dist")

    _años_baj_opc = ["Todos"] + sorted(
        df_bajas["_fecha_baja"].dropna().dt.year.unique().astype(int).tolist(), reverse=True
    )
    _filtro_año = _fc_y.selectbox("Año de baja", _años_baj_opc, key="fil_año_baj")

    _b2b_opc = ["Todos"] + sorted(
        df_bajas["B2B - B2C"].dropna().replace("", pd.NA).dropna().unique().tolist()
    )
    _filtro_b2b = _fc_b2b.selectbox("B2B / B2C", _b2b_opc, key="fil_b2b_baj")

    _df_baj_filt = df_bajas.copy()
    if _filtro_rubro_dist != "Todos":
        _df_baj_filt = _df_baj_filt[_df_baj_filt["Rubro"] == _filtro_rubro_dist]
    if _filtro_año != "Todos":
        _df_baj_filt = _df_baj_filt[_df_baj_filt["_fecha_baja"].dt.year == int(_filtro_año)]
    if _filtro_b2b != "Todos":
        _df_baj_filt = _df_baj_filt[_df_baj_filt["B2B - B2C"] == _filtro_b2b]

    # ── KPIs sobre datos filtrados ────────────────────────────
    _n_baj  = len(_df_baj_filt)
    _ma_baj = _df_baj_filt["Meses activos"].dropna().mean()
    _mb_str = str(int(round(_ma_baj))) if pd.notna(_ma_baj) else "–"
    _n_baj_filt = _n_baj

    # Distribución por intervalos (sobre bajas filtradas)
    _ma_vals_b = _df_baj_filt["Meses activos"].dropna()
    _dist_rows_b = []
    for lbl, lo, hi in _IVLS_ACT:
        n = int(((lo <= _ma_vals_b) & (_ma_vals_b <= hi)).sum())
        _dist_rows_b.append({"Intervalo": lbl, "Bajas": n,
                              "%": f"{round(n / _n_baj_filt * 100)}%" if _n_baj_filt > 0 else "–"})
    _df_dist_b = pd.DataFrame(_dist_rows_b)

    # Top rubros bajas (siempre sobre total sin filtro para el ranking)
    _n_baj_total = len(df_bajas)
    _rubro_counts_b = (
        df_bajas["Rubro"].fillna("").replace("", "Sin rubro").value_counts()
    )
    _top7_b  = _rubro_counts_b.head(7)
    _otros_b = _rubro_counts_b.iloc[7:].sum()
    _rubro_rows_b = [{"Rubro": r, "Bajas": int(n),
                      "%": f"{round(n / _n_baj_total * 100)}%" if _n_baj_total > 0 else "–"}
                     for r, n in _top7_b.items()]
    if _otros_b > 0:
        _rubro_rows_b.append({"Rubro": "Otros", "Bajas": int(_otros_b),
                               "%": f"{round(_otros_b / _n_baj_total * 100)}%" if _n_baj_total > 0 else "–"})
    _df_rubro_b = pd.DataFrame(_rubro_rows_b)

    _bkpi1, _bdist, _brubro = st.columns([0.6, 1, 1.4])
    with _bkpi1:
        st.markdown(
            _kpi("Clientes dados de baja", f"{_n_baj:,}")
            + '<div style="margin-top:12px"></div>'
            + _kpi("Meses activos prom.", _mb_str, "antes de la baja"),
            unsafe_allow_html=True,
        )
    with _bdist:
        st.markdown(_tabla_html(_df_dist_b, "Intervalo"), unsafe_allow_html=True)
    with _brubro:
        st.markdown(_tabla_html(_df_rubro_b, "Rubro"), unsafe_allow_html=True)

    st.markdown('<div style="margin-top:18px"></div>', unsafe_allow_html=True)

    # ── Tabla detalle ─────────────────────────────────────────

    _buscar_b = st.text_input("🔍 Buscar nombre o ID", "", key="buscar_bajas")

    _cols_baj = ["ID CRM", "Nombre", "Fecha de ingreso", "Fecha de baja",
                 "Meses activos", "Tipo de cliente", "Rubro", "B2B - B2C"]
    _show_b   = (
        _df_baj_filt
        .sort_values("_fecha_baja", ascending=False, na_position="last")
        [_cols_baj]
        .rename(columns={
            "ID CRM":           "ID",
            "Fecha de ingreso":  "Fecha ing",
            "Fecha de baja":     "Fecha baja",
            "Meses activos":     "Meses act",
        })
    )

    if _buscar_b:
        _mask = (
            _show_b["Nombre"].str.contains(_buscar_b, case=False, na=False)
            | _show_b["ID"].str.contains(_buscar_b, case=False, na=False)
        )
        _show_b = _show_b[_mask]

    st.markdown(f"**{len(_show_b):,} bajas**")
    st.dataframe(
        _show_b, use_container_width=True, hide_index=True,
        column_config={
            "ID":          st.column_config.TextColumn("ID",          width=None),
            "Nombre":      st.column_config.TextColumn("Nombre",      width="medium"),
            "Fecha ing":   st.column_config.TextColumn("Fecha ing",   width=None),
            "Fecha baja":  st.column_config.TextColumn("Fecha baja",  width=None),
            "Meses act":   st.column_config.NumberColumn(
                "Meses act", format="%.0f", width="small",
                help="(Fecha de baja − Fecha de ingreso) ÷ 30",
            ),
            "Tipo de cliente": st.column_config.TextColumn("Tipo de cliente", width=None),
            "B2B - B2C":   st.column_config.TextColumn("B2B - B2C",   width=None),
        },
    )

    # ══════════════════════════════════════════════════════════
    #  TABLA 1: Cohorte de baja × intervalo de duración
    # ══════════════════════════════════════════════════════════

    st.markdown("#### Bajas por cohorte de baja e intervalo de duración")

    _col_lbl1, _col_rad1, _ = st.columns([1, 2, 6])
    with _col_lbl1:
        st.markdown(
            '<div style="padding-top:8px;font-size:0.85rem;color:#444">Agrupar por:</div>',
            unsafe_allow_html=True,
        )
    with _col_rad1:
        _gran1 = st.radio(
            "", ["Mes", "Trimestre"],
            horizontal=True, key="gran_coh1", label_visibility="collapsed",
        )

    st.markdown(
        '<div style="font-size:0.7rem;color:#999;margin-bottom:4px">'
        "Cohorte = mes en que se dio de baja &nbsp;·&nbsp; "
        "Intervalos: <b>0</b> ≤ 0m &nbsp;·&nbsp; <b>0-3</b> = 1-3m &nbsp;·&nbsp; "
        "<b>+3</b> = 3-6m &nbsp;·&nbsp; <b>+6</b> = 6-12m &nbsp;·&nbsp; <b>+12</b> > 12m &nbsp;·&nbsp; "
        "Desde 2022 · Click en el año para expandir"
        "</div>",
        unsafe_allow_html=True,
    )

    _html_c1, _h_c1 = _cohorte_html(
        df_bajas,
        granularity="trimestre" if _gran1 == "Trimestre" else "mes",
    )
    components.html(_html_c1, height=_h_c1, scrolling=False)

    # ══════════════════════════════════════════════════════════
    #  TABLA 2: Cohorte de VENTA × bajas (Monday)
    # ══════════════════════════════════════════════════════════

    st.markdown(
        '<div style="margin-top:8px"></div>',
        unsafe_allow_html=True,
    )
    st.markdown("#### Bajas por cohorte de venta")

    _col_lbl2, _col_rad2, _ = st.columns([1, 2, 6])
    with _col_lbl2:
        st.markdown(
            '<div style="padding-top:8px;font-size:0.85rem;color:#444">Agrupar por:</div>',
            unsafe_allow_html=True,
        )
    with _col_rad2:
        _gran2 = st.radio(
            "", ["Mes", "Trimestre"],
            horizontal=True, key="gran_coh2", label_visibility="collapsed",
        )

    st.markdown(
        '<div style="font-size:0.7rem;color:#999;margin-bottom:4px">'
        "Cohorte = mes de la venta (BBDD_Ventas) &nbsp;·&nbsp; "
        "Bajas: clientes de ese cohorte dados de baja en Monday &nbsp;·&nbsp; "
        "% sobre ventas: proporción respecto al total de ventas del cohorte"
        "</div>",
        unsafe_allow_html=True,
    )

    if not df_ventas_cs.empty:
        _html_c2, _h_c2 = _cohorte_ventas_html(
            df_bajas,
            df_ventas_cs,
            df_activos,
            granularity="trimestre" if _gran2 == "Trimestre" else "mes",
        )
        components.html(_html_c2, height=_h_c2, scrolling=False)
    else:
        st.info("Sin datos de ventas disponibles.")

    st.markdown(
        '<div style="text-align:right;font-size:0.68rem;color:#94a3b8;margin-top:4px">'
        'Fuente: BBDD_Ventas (cohortes) · Monday.com (bajas) · '
        'Matcheo por ID CRM · Un cliente = primera venta'
        '</div>',
        unsafe_allow_html=True,
    )

    _xlsx_bytes = _export_cohorte_excel(
        df_bajas, df_ventas_cs, df_activos, _gran1, _gran2
    )
    st.download_button(
        label="⬇️ Exportar cohortes a Excel",
        data=_xlsx_bytes,
        file_name=f"cohortes_cs_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

